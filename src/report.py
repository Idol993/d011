import os
import json
import platform
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional
from collections import Counter, defaultdict

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    _cn_fonts = []
    if platform.system() == "Windows":
        _cn_fonts = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
    elif platform.system() == "Darwin":
        _cn_fonts = ["PingFang SC", "Arial Unicode MS"]
    else:
        _cn_fonts = ["Noto Sans CJK SC", "WenQuanYi Micro Hei"]
    matplotlib.rcParams["font.sans-serif"] = _cn_fonts + matplotlib.rcParams.get("font.sans-serif", [])
    matplotlib.rcParams["axes.unicode_minus"] = False
except Exception:
    plt = None

_CN_FONT_PATH = None
if platform.system() == "Windows":
    _windir = os.environ.get("WINDIR", r"C:\Windows")
    _candidates = [
        os.path.join(_windir, "Fonts", "msyh.ttc"),
        os.path.join(_windir, "Fonts", "msyhbd.ttc"),
        os.path.join(_windir, "Fonts", "simhei.ttf"),
    ]
    for _fp in _candidates:
        if os.path.exists(_fp):
            _CN_FONT_PATH = _fp
            break

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    if _CN_FONT_PATH:
        try:
            pdfmetrics.registerFont(TTFont("CNFont", _CN_FONT_PATH))
            _PDF_CN_FONT = "CNFont"
        except Exception:
            _PDF_CN_FONT = None
    else:
        _PDF_CN_FONT = None
except Exception:
    SimpleDocTemplate = None
    _PDF_CN_FONT = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    Workbook = None

from .config import REPORTS_DIR, RELEASE_STATUS_LABELS, SUCCESS_STATUSES
from .logger import get_logger
from . import database as db
from . import notifier

logger = get_logger("report")


def _get_week_range(ref_date: Optional[datetime] = None) -> Tuple[str, str, datetime, datetime]:
    ref = ref_date or datetime.now()
    start = ref - timedelta(days=ref.weekday())
    start = start.replace(hour=0, minute=0, second=0, microsecond=0)
    end = start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    return start.isoformat(timespec="seconds"), end.isoformat(timespec="seconds"), start, end


def collect_weekly_stats(ref_date: Optional[datetime] = None) -> Dict:
    start_str, end_str, start_dt, end_dt = _get_week_range(ref_date)
    filters = {"start_time": start_str, "end_time": end_str}
    releases = db.list_releases(filters)

    release_total = len(releases)

    status_counts = {}
    for status in RELEASE_STATUS_LABELS:
        status_counts[status] = 0
    for r in releases:
        s = r["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    risk_counts = {}
    for r in releases:
        risk_counts[r["risk_level"]] = risk_counts.get(r["risk_level"], 0) + 1

    release_success = len([r for r in releases
                           if r["status"] in SUCCESS_STATUSES
                           and not r["rollback_triggered"]])
    rollback_count = len([r for r in releases if r["rollback_triggered"]])
    release_failed = len([r for r in releases if r["status"] in {"precheck_failed", "rejected", "rolled_back"}])
    release_in_progress = len([r for r in releases if r["status"] in {"pending", "awaiting_approval", "approved", "grayscale"}])

    durations = []
    per_role_all = {}
    for r in releases:
        d = db.get_approval_duration_seconds(r["id"])
        if d is not None:
            durations.append(d)
        rd = db.get_per_role_durations(r["id"])
        for role, dur in rd.items():
            if dur is not None:
                per_role_all.setdefault(role, []).append(dur)
    avg_approval = sum(durations) / len(durations) if durations else 0.0

    per_role_avg = {}
    for role, durs in per_role_all.items():
        per_role_avg[role] = sum(durs) / len(durs)

    daily_stats = {}
    for i in range(7):
        day = (start_dt + timedelta(days=i)).strftime("%Y-%m-%d")
        daily_stats[day] = {"releases": 0, "success": 0, "rollbacks": 0, "failed": 0}
    for r in releases:
        try:
            r_day = r["created_at"][:10]
            if r_day in daily_stats:
                daily_stats[r_day]["releases"] += 1
                if r["status"] in SUCCESS_STATUSES and not r["rollback_triggered"]:
                    daily_stats[r_day]["success"] += 1
                if r["rollback_triggered"]:
                    daily_stats[r_day]["rollbacks"] += 1
                if r["status"] in {"precheck_failed", "rejected", "rolled_back"}:
                    daily_stats[r_day]["failed"] += 1
        except Exception:
            pass

    prev_week_start = (start_dt - timedelta(days=7)).isoformat(timespec="seconds")
    prev_filters = {"start_time": prev_week_start, "end_time": start_str}
    prev_releases = db.list_releases(prev_filters)
    prev_total = len(prev_releases)
    prev_success = len([r for r in prev_releases if r["status"] in SUCCESS_STATUSES and not r["rollback_triggered"]])
    prev_rollback = len([r for r in prev_releases if r["rollback_triggered"]])

    risk_ranking = []
    for risk in ["emergency", "normal"]:
        subset = [r for r in releases if r["risk_level"] == risk]
        total = len(subset)
        success = len([r for r in subset if r["status"] in SUCCESS_STATUSES and not r["rollback_triggered"]])
        rollback = len([r for r in subset if r["rollback_triggered"]])
        failed = len([r for r in subset if r["status"] in {"precheck_failed", "rejected", "rolled_back"}])
        risk_ranking.append({
            "risk_level": risk,
            "total": total,
            "success": success,
            "success_rate": (success / total) if total else 0.0,
            "failed": failed,
            "rollback_count": rollback,
        })
    risk_ranking.sort(key=lambda x: (-x["rollback_count"], -x["failed"]))

    rollback_reasons_counter: Counter = Counter()
    rollback_details_list = []
    for r in releases:
        if r["rollback_triggered"] or r["status"] == "rolled_back":
            rbs = db.list_rollbacks(r["id"])
            for rb in rbs:
                reason = rb.get("reason") or r.get("rollback_reason") or "未填写原因"
                rollback_reasons_counter[reason] += 1
                rollback_details_list.append({
                    "version": r["version"],
                    "release_id": r["id"],
                    "reason": reason,
                    "affected_centers": rb.get("affected_centers") or [],
                    "affected_parcels": rb.get("affected_parcels", 0),
                    "created_at": rb.get("created_at"),
                })
    rollback_reason_top = [
        {"reason": reason, "count": count}
        for reason, count in rollback_reasons_counter.most_common(10)
    ]

    approval_timeout_top = []
    timeout_approvals_all = []
    from src.config import APPROVAL_TIMEOUT_MINUTES
    for r in releases:
        approvals = db.list_approvals(r["id"])
        for a in approvals:
            threshold_min = APPROVAL_TIMEOUT_MINUTES.get(r["risk_level"], 480)
            if a.get("created_at"):
                try:
                    elapsed_min = None
                    if a.get("approved_at"):
                        elapsed_min = (datetime.fromisoformat(a["approved_at"])
                                       - datetime.fromisoformat(a["created_at"])).total_seconds() / 60
                    elif a["status"] == "pending":
                        elapsed_min = (datetime.now()
                                       - datetime.fromisoformat(a["created_at"])).total_seconds() / 60
                    if elapsed_min and elapsed_min > threshold_min:
                        timeout_approvals_all.append({
                            "version": r["version"],
                            "release_id": r["id"],
                            "role": a["role"],
                            "approver": a.get("approver"),
                            "status": a["status"],
                            "elapsed_minutes": round(elapsed_min, 1),
                            "threshold_minutes": threshold_min,
                            "timeout_reminded": bool(a.get("timeout_reminded", 0)),
                        })
                except Exception:
                    pass
    timeout_approvals_all.sort(key=lambda x: -x["elapsed_minutes"])
    approval_timeout_top = timeout_approvals_all[:10]

    per_center_total: defaultdict = defaultdict(int)
    per_center_success: defaultdict = defaultdict(int)
    per_center_rollback: defaultdict = defaultdict(int)
    per_center_failed: defaultdict = defaultdict(int)
    for r in releases:
        rbs = db.list_rollbacks(r["id"])
        rollback_centers = set()
        for rb in rbs:
            for c in rb.get("affected_centers", []):
                rollback_centers.add(c)
        stages = db.list_grayscale_stages(r["id"])
        deployed = set()
        for s in stages:
            for c in s.get("center_ids", []):
                deployed.add(c)
        is_success = r["status"] in SUCCESS_STATUSES and not r["rollback_triggered"]
        is_failed = r["status"] in {"precheck_failed", "rejected", "rolled_back"}
        for c in deployed:
            per_center_total[c] += 1
            if is_success and c not in rollback_centers:
                per_center_success[c] += 1
            if c in rollback_centers:
                per_center_rollback[c] += 1
            if is_failed:
                per_center_failed[c] += 1
        if not deployed and (is_success or is_failed):
            fallback_centers = ["DC001"]
            for c in fallback_centers:
                per_center_total[c] += 1
                if is_success:
                    per_center_success[c] += 1
                if is_failed:
                    per_center_failed[c] += 1

    center_success_rates = []
    all_centers = set(list(per_center_total.keys()))
    from src.config import DISTRIBUTION_CENTERS
    for dc in DISTRIBUTION_CENTERS:
        all_centers.add(dc["id"])
    for cid in sorted(all_centers):
        total = per_center_total[cid]
        succ = per_center_success[cid]
        rb = per_center_rollback[cid]
        fl = per_center_failed[cid]
        center_success_rates.append({
            "center_id": cid,
            "center_name": next((c["name"] for c in DISTRIBUTION_CENTERS if c["id"] == cid), cid),
            "release_total": total,
            "success": succ,
            "success_rate": (succ / total) if total else 0.0,
            "rollback_count": rb,
            "failed_count": fl,
        })
    center_success_rates.sort(key=lambda x: (-x["success_rate"] if x["release_total"] else -1, -x["release_total"]))

    return {
        "week_start": start_str,
        "week_end": end_str,
        "start_dt": start_dt,
        "end_dt": end_dt,
        "release_total": release_total,
        "release_success": release_success,
        "release_success_rate": (release_success / release_total) if release_total else 0.0,
        "release_failed": release_failed,
        "release_in_progress": release_in_progress,
        "rollback_count": rollback_count,
        "avg_approval_seconds": avg_approval,
        "per_role_avg": per_role_avg,
        "releases": releases,
        "status_counts": status_counts,
        "risk_counts": risk_counts,
        "daily_stats": daily_stats,
        "prev_week": {
            "total": prev_total,
            "success": prev_success,
            "success_rate": (prev_success / prev_total) if prev_total else 0.0,
            "rollback_count": prev_rollback,
        },
        "json_analytics": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "week_range": {"start": start_str, "end": end_str},
            "risk_ranking": risk_ranking,
            "rollback_reason_top": rollback_reason_top,
            "rollback_details": rollback_details_list,
            "approval_timeout_top": approval_timeout_top,
            "center_success_rates": center_success_rates,
            "status_distribution": [
                {"status": k, "status_cn": RELEASE_STATUS_LABELS.get(k, k), "count": v}
                for k, v in status_counts.items() if v > 0
            ],
            "risk_distribution": [
                {"risk_level": k, "count": v} for k, v in risk_counts.items()
            ],
            "core_metrics": {
                "release_total": release_total,
                "release_success": release_success,
                "release_success_rate": (release_success / release_total) if release_total else 0.0,
                "release_failed": release_failed,
                "release_in_progress": release_in_progress,
                "rollback_count": rollback_count,
                "avg_approval_minutes": avg_approval / 60,
                "per_role_avg_minutes": {k: v / 60 for k, v in per_role_avg.items()},
            },
            "weekly_comparison": {
                "current": {
                    "total": release_total,
                    "success": release_success,
                    "success_rate": (release_success / release_total) if release_total else 0.0,
                    "rollback_count": rollback_count,
                },
                "previous": {
                    "total": prev_total,
                    "success": prev_success,
                    "success_rate": (prev_success / prev_total) if prev_total else 0.0,
                    "rollback_count": prev_rollback,
                },
                "delta": {
                    "total_delta": release_total - prev_total,
                    "success_delta": release_success - prev_success,
                    "success_rate_delta_pct": (
                        ((release_success / release_total) if release_total else 0.0)
                        - ((prev_success / prev_total) if prev_total else 0.0)
                    ) * 100,
                    "rollback_delta": rollback_count - prev_rollback,
                },
            },
        },
    }


def _make_json_summary(stats: Dict, out_path: str) -> Optional[str]:
    try:
        analytics = stats["json_analytics"]
        brief_releases = []
        for r in stats["releases"]:
            brief_releases.append({
                "release_id": r["id"],
                "version": r["version"],
                "risk_level": r["risk_level"],
                "status": r["status"],
                "status_cn": RELEASE_STATUS_LABELS.get(r["status"], r["status"]),
                "submitter": r["submitter"],
                "rollback_triggered": bool(r["rollback_triggered"]),
                "emergency_urgent": bool(r.get("emergency_urgent", 0)),
                "created_at": r["created_at"],
                "updated_at": r["updated_at"],
            })
        output = {
            **analytics,
            "releases": brief_releases,
        }
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"周报 JSON 汇总已生成: {out_path}")
        return out_path
    except Exception as e:
        logger.error(f"生成 JSON 汇总失败: {e}")
        return None


def _make_chart(stats: Dict, out_path: str) -> Optional[str]:
    if plt is None:
        return None
    try:
        days = list(stats["daily_stats"].keys())
        success = [s["success"] for s in stats["daily_stats"].values()]
        failed = [s["failed"] for s in stats["daily_stats"].values()]
        rollbacks = [s["rollbacks"] for s in stats["daily_stats"].values()]
        in_prog = [s["releases"] - s["success"] - s["failed"] - s["rollbacks"]
                   for s in stats["daily_stats"].values()]

        fig, axes = plt.subplots(1, 3, figsize=(16, 4.5))
        x = range(len(days))
        short_days = [d[5:] for d in days]

        b1 = axes[0].bar(x, success, label="成功", color="#55A868", alpha=0.85)
        b2 = axes[0].bar(x, in_prog, bottom=success, label="进行中", color="#4C72B0", alpha=0.85)
        b3 = axes[0].bar(x, failed, bottom=[s+i for s, i in zip(success, in_prog)],
                         label="失败", color="#C44E52", alpha=0.85)
        b4 = axes[0].bar(x, rollbacks,
                         bottom=[s+i+f for s, i, f in zip(success, in_prog, failed)],
                         label="回滚", color="#DD8452", alpha=0.85)
        axes[0].set_xticks(list(x))
        axes[0].set_xticklabels(short_days, rotation=30)
        axes[0].set_title("每日发布状态分布")
        axes[0].legend(fontsize=8)
        axes[0].grid(axis="y", linestyle="--", alpha=0.4)

        risk_labels = list(stats["risk_counts"].keys())
        risk_values = list(stats["risk_counts"].values())
        if not risk_labels:
            risk_labels, risk_values = ["无数据"], [1]
        risk_colors = ["#4C72B0", "#C44E52", "#55A868"]
        axes[1].pie(risk_values, labels=risk_labels, autopct="%1.1f%%",
                    colors=risk_colors[:len(risk_labels)], startangle=90)
        axes[1].set_title("风险级别占比")

        pw = stats["prev_week"]
        cmp_labels = ["本周", "上周"]
        cmp_success = [stats["release_success_rate"] * 100, pw["success_rate"] * 100]
        cmp_rollback = [stats["rollback_count"], pw["rollback_count"]]
        ax3a = axes[2]
        ax3b = ax3a.twinx()
        ax3a.bar([0, 1], cmp_success, width=0.35, label="成功率(%)", color="#55A868", alpha=0.85)
        ax3b.bar([0.4, 1.4], cmp_rollback, width=0.35, label="回滚次数", color="#C44E52", alpha=0.85)
        ax3a.set_xticks([0.2, 1.2])
        ax3a.set_xticklabels(cmp_labels)
        ax3a.set_ylabel("成功率(%)")
        ax3b.set_ylabel("回滚次数")
        ax3a.set_title("周环比对比")
        ax3a.legend(loc="upper left", fontsize=7)
        ax3b.legend(loc="upper right", fontsize=7)
        ax3a.grid(axis="y", linestyle="--", alpha=0.3)

        plt.tight_layout()
        plt.savefig(out_path, dpi=120)
        plt.close(fig)
        return out_path
    except Exception as e:
        logger.error(f"生成图表失败: {e}")
        return None


def _make_pdf(stats: Dict, chart_path: Optional[str], out_path: str) -> Optional[str]:
    if SimpleDocTemplate is None:
        return None
    try:
        doc = SimpleDocTemplate(out_path, pagesize=A4)
        styles = getSampleStyleSheet()
        fn = _PDF_CN_FONT or "Helvetica"
        fnb = _PDF_CN_FONT or "Helvetica-Bold"
        title_style = ParagraphStyle("TitleCN", parent=styles["Title"], fontName=fnb, fontSize=18, leading=24)
        h2 = ParagraphStyle("H2CN", parent=styles["Heading2"], fontName=fnb, fontSize=13, leading=18)
        body = ParagraphStyle("BodyCN", parent=styles["BodyText"], fontName=fn, fontSize=10, leading=14)
        small = ParagraphStyle("SmallCN", parent=styles["BodyText"], fontName=fn, fontSize=8, leading=10)

        story = []
        story.append(Paragraph("快递快运分拨系统 - 发布与回滚周报", title_style))
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(f"统计周期: {stats['week_start']} ~ {stats['week_end']}", body))
        story.append(Spacer(1, 0.4 * cm))

        story.append(Paragraph("一、核心指标", h2))
        story.append(Spacer(1, 0.2 * cm))
        rate = stats["release_success_rate"] * 100
        avg_min = stats["avg_approval_seconds"] / 60
        pw = stats["prev_week"]
        pw_rate = pw["success_rate"] * 100
        rate_diff = rate - pw_rate
        rate_trend = f"({'+' if rate_diff >= 0 else ''}{rate_diff:.1f}pp vs上周)" if stats["release_total"] and pw["total"] else ""

        summary_data = [
            ["指标", "本周", "上周", "变化"],
            ["发布总数", str(stats["release_total"]), str(pw["total"]), ""],
            ["成功发布数", str(stats["release_success"]), str(pw["success"]), ""],
            ["发布成功率", f"{rate:.2f}%", f"{pw_rate:.2f}%", rate_trend],
            ["回滚次数", str(stats["rollback_count"]), str(pw["rollback_count"]), ""],
            ["失败/驳回", str(stats["release_failed"]), "", ""],
            ["进行中", str(stats["release_in_progress"]), "", ""],
            ["平均审批时长", f"{avg_min:.2f} 分钟", "", ""],
        ]
        t = Table(summary_data, colWidths=[5 * cm, 3.5 * cm, 3.5 * cm, 4 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4C72B0")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), fnb),
            ("FONTNAME", (0, 1), (-1, -1), fn),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4 * cm))

        story.append(Paragraph("二、各状态分布", h2))
        story.append(Spacer(1, 0.2 * cm))
        status_data = [["状态", "数量", "占比"]]
        for status, count in stats["status_counts"].items():
            label = RELEASE_STATUS_LABELS.get(status, status)
            pct = f"{count/stats['release_total']*100:.1f}%" if stats["release_total"] else "0%"
            status_data.append([label, str(count), pct])
        st = Table(status_data, colWidths=[5 * cm, 4 * cm, 4 * cm])
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#55A868")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), fnb),
            ("FONTNAME", (0, 1), (-1, -1), fn),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(st)
        story.append(Spacer(1, 0.4 * cm))

        if stats["per_role_avg"]:
            story.append(Paragraph("三、各角色平均审批耗时", h2))
            story.append(Spacer(1, 0.2 * cm))
            from .config import STAKEHOLDERS
            role_data = [["角色", "平均耗时"]]
            for role, avg_s in stats["per_role_avg"].items():
                name = STAKEHOLDERS.get(role, {}).get("name", role)
                role_data.append([f"{name}({role})", f"{avg_s/60:.1f} 分钟"])
            rt = Table(role_data, colWidths=[8 * cm, 6 * cm])
            rt.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8172B2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), fnb),
                ("FONTNAME", (0, 1), (-1, -1), fn),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            story.append(rt)
            story.append(Spacer(1, 0.4 * cm))

        chart_section = "四" if stats["per_role_avg"] else "三"
        detail_section = "五" if stats["per_role_avg"] else "四"
        story.append(Paragraph(f"{chart_section}、趋势图表", h2))
        story.append(Spacer(1, 0.2 * cm))
        if chart_path and os.path.exists(chart_path):
            story.append(Image(chart_path, width=17 * cm, height=5 * cm))
        else:
            story.append(Paragraph("(图表生成失败)", body))
        story.append(Spacer(1, 0.4 * cm))

        story.append(Paragraph(f"{detail_section}、发布明细", h2))
        story.append(Spacer(1, 0.2 * cm))
        detail_header = ["版本", "风险", "状态", "提交者", "加急", "提交时间"]
        detail_data = [detail_header]
        for r in stats["releases"][:20]:
            detail_data.append([
                r["version"],
                r["risk_level"],
                RELEASE_STATUS_LABELS.get(r["status"], r["status"]),
                r["submitter"],
                "是" if r.get("emergency_urgent") else "否",
                r["created_at"][:19],
            ])
        if len(detail_data) == 1:
            detail_data.append(["(无数据)", "", "", "", "", ""])
        col_w = [3 * cm, 2 * cm, 2.8 * cm, 2.5 * cm, 1.5 * cm, 4 * cm]
        dt = Table(detail_data, colWidths=col_w)
        dt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#55A868")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), fnb),
            ("FONTNAME", (0, 1), (-1, -1), fn),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(dt)

        doc.build(story)
        return out_path
    except Exception as e:
        logger.error(f"生成PDF失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def _make_excel(stats: Dict, out_path: str) -> Optional[str]:
    if Workbook is None:
        return None
    try:
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4C72B0")
        center = Alignment(horizontal="center", vertical="center")

        ws1 = wb.active
        ws1.title = "核心指标"
        ws1.append(["指标", "本周", "上周"])
        for c in ws1[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        rate = stats["release_success_rate"] * 100
        avg_min = stats["avg_approval_seconds"] / 60
        pw = stats["prev_week"]
        pw_rate = pw["success_rate"] * 100
        rows = [
            ["发布总数", stats["release_total"], pw["total"]],
            ["成功发布数(已发布且未回滚)", stats["release_success"], pw["success"]],
            ["发布成功率", f"{rate:.2f}%", f"{pw_rate:.2f}%"],
            ["回滚次数", stats["rollback_count"], pw["rollback_count"]],
            ["失败/驳回数", stats["release_failed"], ""],
            ["进行中", stats["release_in_progress"], ""],
            ["平均审批时长(分钟)", f"{avg_min:.2f}", ""],
            ["统计开始时间", stats["week_start"], ""],
            ["统计结束时间", stats["week_end"], ""],
        ]
        for r in rows:
            ws1.append(r)
        for row in ws1.iter_rows():
            for c in row:
                c.alignment = center
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 20

        ws_status = wb.create_sheet("状态分布")
        ws_status.append(["状态", "数量", "占比"])
        for c in ws_status[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for status, count in stats["status_counts"].items():
            label = RELEASE_STATUS_LABELS.get(status, status)
            pct = f"{count/stats['release_total']*100:.1f}%" if stats["release_total"] else "0%"
            ws_status.append([label, count, pct])
        ws_status.column_dimensions["A"].width = 18
        ws_status.column_dimensions["B"].width = 10
        ws_status.column_dimensions["C"].width = 12

        ws_risk = wb.create_sheet("风险级别")
        ws_risk.append(["风险级别", "数量", "占比"])
        for c in ws_risk[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for risk, count in stats["risk_counts"].items():
            pct = f"{count/stats['release_total']*100:.1f}%" if stats["release_total"] else "0%"
            ws_risk.append([risk, count, pct])
        ws_risk.column_dimensions["A"].width = 15
        ws_risk.column_dimensions["B"].width = 10
        ws_risk.column_dimensions["C"].width = 12

        ws_role = wb.create_sheet("角色审批耗时")
        ws_role.append(["角色", "平均耗时(分钟)"])
        for c in ws_role[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        from .config import STAKEHOLDERS
        for role, avg_s in stats["per_role_avg"].items():
            name = STAKEHOLDERS.get(role, {}).get("name", role)
            ws_role.append([f"{name}({role})", f"{avg_s/60:.1f}"])
        ws_role.column_dimensions["A"].width = 22
        ws_role.column_dimensions["B"].width = 18

        ws2 = wb.create_sheet("每日统计")
        ws2.append(["日期", "发布数", "成功", "失败", "回滚"])
        for c in ws2[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for day, s in stats["daily_stats"].items():
            ws2.append([day, s["releases"], s["success"], s["failed"], s["rollbacks"]])
        ws2.column_dimensions["A"].width = 15
        for col in "BCDE":
            ws2.column_dimensions[col].width = 10

        ws3 = wb.create_sheet("发布明细")
        headers = ["版本号", "风险级别", "状态", "提交人", "描述",
                   "稳定版本", "是否加急", "是否回滚", "回滚原因", "创建时间", "更新时间"]
        ws3.append(headers)
        for c in ws3[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for r in stats["releases"]:
            ws3.append([
                r["version"], r["risk_level"],
                RELEASE_STATUS_LABELS.get(r["status"], r["status"]),
                r["submitter"],
                r.get("description") or "",
                r.get("stable_version") or "",
                "是" if r.get("emergency_urgent") else "否",
                "是" if r["rollback_triggered"] else "否",
                r.get("rollback_reason") or "",
                r["created_at"], r["updated_at"],
            ])
        for i, w_ in enumerate([18, 12, 16, 12, 30, 15, 10, 10, 30, 22, 22], start=1):
            ws3.column_dimensions[chr(64 + i)].width = w_

        ws4 = wb.create_sheet("回滚明细")
        rb_headers = ["回滚ID", "发布ID", "版本", "影响分拨", "异常件量",
                      "原因", "恢复版本", "状态", "创建时间", "完成时间"]
        ws4.append(rb_headers)
        for c in ws4[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        rollbacks = db.list_rollbacks()
        for rb in rollbacks:
            created = rb["created_at"] or ""
            if stats["week_start"] <= created <= stats["week_end"]:
                rel = db.get_release(rb["release_id"])
                ws4.append([
                    rb["id"], rb["release_id"],
                    rel["version"] if rel else "",
                    ", ".join(rb["affected_centers"]),
                    rb["affected_parcels"], rb["reason"],
                    rb["rolled_back_version"], rb["status"],
                    rb["created_at"], rb.get("completed_at") or "",
                ])
        for i, w_ in enumerate([10, 10, 18, 28, 12, 35, 18, 14, 22, 22], start=1):
            ws4.column_dimensions[chr(64 + i)].width = w_

        wb.save(out_path)
        return out_path
    except Exception as e:
        logger.error(f"生成Excel失败: {e}")
        return None


def generate_weekly_report(ref_date: Optional[datetime] = None) -> Dict:
    stats = collect_weekly_stats(ref_date)
    tag = stats["start_dt"].strftime("%Y%m%d")
    chart_path = os.path.join(REPORTS_DIR, f"weekly_chart_{tag}.png")
    pdf_path = os.path.join(REPORTS_DIR, f"weekly_report_{tag}.pdf")
    excel_path = os.path.join(REPORTS_DIR, f"weekly_report_{tag}.xlsx")
    json_path = os.path.join(REPORTS_DIR, f"weekly_summary_{tag}.json")

    chart = _make_chart(stats, chart_path)
    pdf = _make_pdf(stats, chart, pdf_path)
    xls = _make_excel(stats, excel_path)
    jsn = _make_json_summary(stats, json_path)

    report_id = db.insert_weekly_report(
        stats["week_start"], stats["week_end"],
        stats["release_total"], stats["release_success"],
        stats["rollback_count"], stats["avg_approval_seconds"],
        pdf, xls, jsn,
    )
    if pdf and xls:
        notifier.notify_weekly_report_ready(report_id, pdf, xls)

    return {
        "report_id": report_id,
        "stats": stats,
        "pdf": pdf,
        "excel": xls,
        "json": jsn,
        "chart": chart,
    }
