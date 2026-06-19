import os
import json
import csv
from datetime import datetime
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except Exception:
    Workbook = None

from .config import REPORTS_DIR, DISTRIBUTION_CENTERS, RELEASE_STATUS_LABELS, STAKEHOLDERS
from .logger import get_logger
from . import database as db

logger = get_logger("history")


def search_releases(version: Optional[str] = None,
                    status: Optional[str] = None,
                    risk_level: Optional[str] = None,
                    start_time: Optional[str] = None,
                    end_time: Optional[str] = None,
                    center_id: Optional[str] = None) -> List[Dict]:
    filters: Dict = {}
    if version:
        filters["version"] = version
    if status:
        filters["status"] = status
    if risk_level:
        filters["risk_level"] = risk_level
    if start_time:
        filters["start_time"] = start_time
    if end_time:
        filters["end_time"] = end_time

    results = db.list_releases(filters)

    if center_id:
        filtered = []
        for r in results:
            stages = db.list_grayscale_stages(r["id"])
            for s in stages:
                if center_id in s.get("center_ids", []):
                    filtered.append(r)
                    break
        results = filtered

    return results


def _build_detail(release: Dict) -> Dict:
    prechecks = db.list_prechecks(release["id"])
    approvals = db.list_approvals(release["id"])
    stages = db.list_grayscale_stages(release["id"])
    rollbacks = db.list_rollbacks(release["id"])
    role_durations = db.get_per_role_durations(release["id"])
    for a in approvals:
        dur = a.get("duration_seconds")
        a["duration_str"] = f"{dur:.0f}s ({dur/60:.1f}min)" if dur else "-"
        a["timeout_reminded"] = bool(a.get("timeout_reminded", 0))
    return {
        "release": release,
        "prechecks": prechecks,
        "approvals": approvals,
        "grayscale_stages": stages,
        "rollbacks": rollbacks,
        "role_durations": role_durations,
    }


def get_release_detail(release_id: int) -> Optional[Dict]:
    release = db.get_release(release_id)
    if not release:
        return None
    return _build_detail(release)


def get_release_detail_by_version(version: str) -> Optional[Dict]:
    release = db.get_release_by_version(version)
    if not release:
        return None
    return _build_detail(release)


def export_releases(releases: List[Dict], fmt: str = "excel",
                    filename_prefix: str = "release_export") -> Optional[str]:
    if not releases:
        logger.warning("无数据可导出")
        return None

    tag = datetime.now().strftime("%Y%m%d_%H%M%S")
    fmt = fmt.lower()

    if fmt == "json":
        path = os.path.join(REPORTS_DIR, f"{filename_prefix}_{tag}.json")
        full_data = []
        for r in releases:
            full_data.append(_build_detail(r))
        with open(path, "w", encoding="utf-8") as f:
            json.dump(full_data, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"已导出 JSON: {path}")
        return path

    if fmt == "csv":
        path = os.path.join(REPORTS_DIR, f"{filename_prefix}_{tag}.csv")
        headers = ["ID", "版本号", "风险级别", "状态", "提交人", "描述",
                   "稳定版本", "是否加急", "是否回滚", "回滚原因", "创建时间", "更新时间"]
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in releases:
                w.writerow([
                    r["id"], r["version"], r["risk_level"],
                    RELEASE_STATUS_LABELS.get(r["status"], r["status"]),
                    r["submitter"], r.get("description") or "",
                    r.get("stable_version") or "",
                    "是" if r.get("emergency_urgent") else "否",
                    "是" if r["rollback_triggered"] else "否",
                    r.get("rollback_reason") or "",
                    r["created_at"], r["updated_at"],
                ])
        logger.info(f"已导出 CSV: {path}")
        return path

    if fmt in ("excel", "xlsx"):
        if Workbook is None:
            logger.error("openpyxl 未安装，无法导出Excel")
            return None
        path = os.path.join(REPORTS_DIR, f"{filename_prefix}_{tag}.xlsx")
        wb = Workbook()

        ws = wb.active
        ws.title = "发布记录"
        headers = ["ID", "版本号", "风险级别", "状态", "提交人", "描述",
                   "稳定版本", "是否加急", "是否回滚", "回滚原因", "创建时间", "更新时间"]
        ws.append(headers)
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="4C72B0")
        center = Alignment(horizontal="center", vertical="center")
        for c in ws[1]:
            c.font = hf
            c.fill = hfill
            c.alignment = center
        for r in releases:
            ws.append([
                r["id"], r["version"], r["risk_level"],
                RELEASE_STATUS_LABELS.get(r["status"], r["status"]),
                r["submitter"], r.get("description") or "",
                r.get("stable_version") or "",
                "是" if r.get("emergency_urgent") else "否",
                "是" if r["rollback_triggered"] else "否",
                r.get("rollback_reason") or "",
                r["created_at"], r["updated_at"],
            ])
        widths = [6, 18, 12, 16, 12, 30, 15, 10, 10, 30, 22, 22]
        for i, w_ in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w_

        ws2 = wb.create_sheet("审批记录")
        ws2.append(["发布ID", "版本", "角色", "审批人", "状态", "审批意见",
                    "耗时", "超时已提醒", "审批时间", "创建时间"])
        for c in ws2[1]:
            c.font = hf
            c.fill = hfill
            c.alignment = center
        for r in releases:
            for a in db.list_approvals(r["id"]):
                dur = a.get("duration_seconds")
                dur_str = f"{dur/60:.1f}min" if dur else "-"
                ws2.append([
                    r["id"], r["version"], a["role"],
                    a.get("approver") or "",
                    a["status"],
                    a.get("comment") or "",
                    dur_str,
                    "是" if a.get("timeout_reminded") else "否",
                    a.get("approved_at") or "",
                    a.get("created_at") or "",
                ])
        for i, w_ in enumerate([10, 18, 14, 14, 12, 30, 14, 12, 22, 22], start=1):
            ws2.column_dimensions[chr(64 + i)].width = w_

        ws3 = wb.create_sheet("回滚记录")
        ws3.append(["回滚ID", "发布ID", "版本", "影响分拨", "异常件量",
                    "原因", "恢复版本", "状态", "创建时间", "完成时间"])
        for c in ws3[1]:
            c.font = hf
            c.fill = hfill
            c.alignment = center
        for r in releases:
            for rb in db.list_rollbacks(r["id"]):
                ws3.append([
                    rb["id"], r["id"], r["version"],
                    ", ".join(rb["affected_centers"]),
                    rb["affected_parcels"], rb["reason"],
                    rb["rolled_back_version"], rb["status"],
                    rb["created_at"], rb.get("completed_at") or "",
                ])
        for i, w_ in enumerate([10, 10, 18, 28, 12, 35, 18, 14, 22, 22], start=1):
            ws3.column_dimensions[chr(64 + i)].width = w_

        wb.save(path)
        logger.info(f"已导出 Excel: {path}")
        return path

    logger.error(f"不支持的导出格式: {fmt}")
    return None
